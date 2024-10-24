import tkinter as tk
from pprint import pprint
from tkinter import messagebox, filedialog, Menu, ttk
import sqlite3

import pandas as pd
import xlrd
import openpyxl

# Tạo cơ sở dữ liệu
def init_db():
    conn = sqlite3.connect('student_management.db')
    cursor = conn.cursor()
    # Tạo bảng students
    cursor.executescript('''
        CREATE TABLE IF NOT EXISTS students (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        MSSV TEXT NOT NULL,
        HoDem TEXT NOT NULL,
        Ten TEXT NOT NULL
    );

        CREATE TABLE IF NOT EXISTS courses (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        MonHoc TEXT NOT NULL,
        Dot TEXT NOT NULL,
        Lop TEXT NOT NULL
    );

        CREATE TABLE IF NOT EXISTS student_courses (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        StudentID INTEGER,
        CourseID INTEGER,
        VangCoPhep INTEGER NOT NULL DEFAULT 0,
        VangKhongPhep INTEGER NOT NULL DEFAULT 0,
        TyLeVang DOUBLE NOT NULL DEFAULT 0,
        FOREIGN KEY (StudentID) REFERENCES students(ID),
        FOREIGN KEY (CourseID) REFERENCES courses(ID)
    );

        CREATE TABLE IF NOT EXISTS absences (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        StudentCourseID INTEGER,
        NgayNghi DATE,
        CoPhep BOOLEAN NOT NULL DEFAULT 0,
        FOREIGN KEY (StudentCourseID) REFERENCES student_courses(ID)
    );

        CREATE TABLE IF NOT EXISTS report_statuses (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        CourseID INTEGER,
        SubmissionDate DATE,
        SubmissionTime TIME, 
        Status TEXT NOT NULL, 
        Email TEXT NOT NULL,    
        FOREIGN KEY (CourseID) REFERENCES courses(ID)
    );
    
        CREATE TABLE IF NOT EXISTS questions (
        ID INTEGER PRIMARY KEY AUTOINCREMENT,
        StudentID INTEGER,
        QuestionText TEXT NOT NULL,
        SubmissionDate DATE NOT NULL,
        SubmissionTime TIME NOT NULL,
        Status TEXT NOT NULL DEFAULT 'Pending',  -- 'Pending', 'Resolved', 'Closed'
        Email TEXT NOT NULL,
        FOREIGN KEY (StudentID) REFERENCES students(ID)
    );

    ''')
    conn.commit()
    conn.close()

# Ứng dụng chính
class StudentManagementApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Quản lý sinh viên")
        self.geometry("1200x500")

        self.init_menu()
        self.init_login_frame()

    def init_menu(self):
        menu_bar = Menu(self)
        self.config(menu=menu_bar)

        # Menu Chức năng
        function_menu = Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Chức năng", menu=function_menu)
        function_menu.add_command(label="Nhập dữ liệu", command=self.import_data)
        function_menu.add_command(label="Thoát", command=self.quit)

    def init_login_frame(self):
        self.login_frame = tk.Frame(self)
        self.login_frame.pack(pady=20)

        tk.Label(self.login_frame, text="Tên đăng nhập:").grid(row=0, column=0)
        self.username_entry = tk.Entry(self.login_frame)
        self.username_entry.grid(row=0, column=1)

        tk.Label(self.login_frame, text="Mật khẩu:").grid(row=1, column=0)
        self.password_entry = tk.Entry(self.login_frame, show='*')
        self.password_entry.grid(row=1, column=1)

        login_button = tk.Button(self.login_frame, text="Đăng nhập", command=self.login)
        login_button.grid(row=2, columnspan=2)

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        # Kiểm tra thông tin đăng nhập
        if username == "admin" and password == "admin":
            self.login_frame.pack_forget()  # Ẩn khung đăng nhập
            self.init_main_frame()
        else:
            messagebox.showerror("Lỗi", "Tên đăng nhập hoặc mật khẩu không đúng.")

    def get_student_details(self, mssv):
      conn = sqlite3.connect("student_management.db")
      cursor = conn.cursor()

      # Thực hiện truy vấn SQL
      query = '''
      SELECT 
          c.MonHoc, 
          sc.VangCoPhep, 
          sc.VangKhongPhep, 
          sc.TyLeVang, 
          a.NgayNghi, 
          a.CoPhep,
          c.Lop
      FROM 
          students s
      JOIN 
          student_courses sc ON s.ID = sc.StudentID
      JOIN 
          courses c ON sc.CourseID = c.ID
      LEFT JOIN 
          absences a ON sc.ID = a.StudentCourseID
      WHERE 
          s.MSSV = ?;
      '''

      cursor.execute(query, (mssv,))
      results = cursor.fetchall()


      student_details = {
          "MonHocs": {}
      }

      for row in results:
          mon_hoc = row[0]
          if mon_hoc not in student_details["MonHocs"]:
              student_details["MonHocs"][mon_hoc] = {
                  "VangCoPhep": row[1],
                  "VangKhongPhep": row[2],
                  "TyLeVang": row[3],
                  "Lop": row[6],
                  "NgayNghi": []
              }

          if row[4]:
              student_details["MonHocs"][mon_hoc]["NgayNghi"].append({
                  "Ngay": row[4],
                  "CoPhep": row[5]
              })

      conn.close()
      #pprint(student_details)
      return student_details

    def init_main_frame(self):
        self.main_frame = tk.Frame(self)
        self.main_frame.pack(pady=20)

        # Khung tìm kiếm
        search_frame = tk.Frame(self.main_frame)
        search_frame.pack(pady=10)

        tk.Label(search_frame, text="Mã sinh viên:").grid(row=0, column=0)
        self.search_id_entry = tk.Entry(search_frame)
        self.search_id_entry.grid(row=0, column=1)

        tk.Label(search_frame, text="Tên sinh viên:").grid(row=0, column=2)
        self.search_name_entry = tk.Entry(search_frame)
        self.search_name_entry.grid(row=0, column=3)

        search_button = tk.Button(search_frame, text="Tìm kiếm", command=self.search_student)
        search_button.grid(row=0, column=4)

        # Combobox chọn Lớp
        tk.Label(search_frame, text="Lớp:").grid(row=1, column=0)
        self.class_combobox = ttk.Combobox(search_frame)
        self.class_combobox.grid(row=1, column=1)

        # Combobox chọn Tên môn học
        tk.Label(search_frame, text="Tên môn học:").grid(row=1, column=2)
        self.course_combobox = ttk.Combobox(search_frame)
        self.course_combobox.grid(row=1, column=3)

        # Combobox sắp xếp theo Tổng buổi vắng
        tk.Label(search_frame, text="Sắp xếp theo:").grid(row=1, column=4)
        self.sort_combobox = ttk.Combobox(search_frame, values=["Họ tên", "Tổng buổi vắng"])
        self.sort_combobox.grid(row=1, column=5)

        sort_button = tk.Button(search_frame, text="Sắp xếp", command=self.sort_students)
        sort_button.grid(row=1, column=6)

        # Tạo bảng hiển thị sinh viên
        self.tree = ttk.Treeview(self.main_frame, columns=("Mã sinh viên", "Họ đệm", "Tên"), show='headings')
        
        # Đặt kích thước cột
        self.tree.column("Mã sinh viên", width=100)
        self.tree.column("Họ đệm", width=100)
        self.tree.column("Tên", width=100)
        
        
        # Đặt tiêu đề cột
        self.tree.heading("Mã sinh viên", text="Mã sinh viên")
        self.tree.heading("Họ đệm", text="Họ đệm")
        self.tree.heading("Tên", text="Tên")
        
        
        self.tree.pack(expand=True, fill='both')
        #self.tree.bind("<Double-1>", self.show_student_details)
        self.tree.bind("<Double-1>", lambda event: self.show_student_details(event))
        # Tải danh sách sinh viên
        self.load_students()

        # Nút thêm sinh viên
        add_button = tk.Button(self.main_frame, text="Thêm sinh viên", command=self.show_add_student_form)
        add_button.pack(pady=5)

        # Nút xóa sinh viên
        delete_button = tk.Button(self.main_frame, text="Xóa sinh viên", command=self.delete_student)
        delete_button.pack(pady=5)

        # Tải dữ liệu cho combobox
        self.load_combobox_values()

    def load_combobox_values(self):
        # Truy vấn danh sách lớp từ cơ sở dữ liệu
        self.conn = sqlite3.connect('student_management.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT DISTINCT Lop FROM courses")
        classes = [row[0] for row in self.cursor.fetchall()]
        #print("classes: " + str(classes))  # In ra để kiểm tra

        # Gán danh sách lớp vào class_combobox
        self.class_combobox['values'] = classes

        # Gán giá trị mặc định cho class_combobox (nếu cần)
        if classes:
            self.class_combobox.set(classes[0])  # Set giá trị đầu tiên trong danh sách làm mặc định

        # Truy vấn danh sách tên môn học từ cơ sở dữ liệu
        self.cursor.execute("SELECT DISTINCT MonHoc FROM courses")
        courses = [row[0] for row in self.cursor.fetchall()]
        #print("courses: " + str(courses))  # In ra để kiểm tra

        # Gán danh sách môn học vào course_combobox
        self.course_combobox['values'] = courses

        # Gán giá trị mặc định cho course_combobox (nếu cần)
        if courses:
            self.course_combobox.set(courses[0])  # Set giá trị đầu tiên trong danh sách làm mặc định



    def sort_students(self):
        sort_option = self.sort_combobox.get()
        self.conn = sqlite3.connect('student_management.db')
        self.cursor = self.conn.cursor()
        print("sort student")
        if sort_option == "Họ tên":
            query = '''
            SELECT s.MSSV, s.HoDem, s.Ten, c.Lop, c.MonHoc, 
                   (sc.VangCoPhep + sc.VangKhongPhep) AS TongBuoiVang
            FROM students s
            JOIN student_courses sc ON s.ID = sc.StudentID
            JOIN courses c ON sc.CourseID = c.ID
            ORDER BY s.HoDem, s.Ten
            '''
        elif sort_option == "Tổng buổi vắng":
            query = '''
            SELECT s.MSSV, s.HoDem, s.Ten, c.Lop, c.MonHoc, 
                   (sc.VangCoPhep + sc.VangKhongPhep) AS TongBuoiVang
            FROM students s
            JOIN student_courses sc ON s.ID = sc.StudentID
            JOIN courses c ON sc.CourseID = c.ID
            ORDER BY TongBuoiVang DESC
            '''
        elif sort_option == "Lớp":
            query = '''
            SELECT s.MSSV, s.HoDem, s.Ten, c.Lop, c.MonHoc, 
                   (sc.VangCoPhep + sc.VangKhongPhep) AS TongBuoiVang
            FROM students s
            JOIN student_courses sc ON s.ID = sc.StudentID
            JOIN courses c ON sc.CourseID = c.ID
            ORDER BY c.Lop
            '''
        elif sort_option == "Tên môn học":
            query = '''
                        SELECT s.MSSV, s.HoDem, s.Ten, c.Lop, c.MonHoc, 
                               (sc.VangCoPhep + sc.VangKhongPhep) AS TongBuoiVang
                        FROM students s
                        JOIN student_courses sc ON s.ID = sc.StudentID
                        JOIN courses c ON sc.CourseID = c.ID
                        ORDER BY c.MonHoc
                        '''
        else:
            return

        self.cursor.execute(query)
        rows = self.cursor.fetchall()

        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in rows:
            self.tree.insert("", "end", values=row)
    def show_student_details(self, event=None):
        selected_item = self.tree.selection()[0]
        mssv = self.tree.item(selected_item)['values'][0]

        student_details = self.get_student_details(mssv)

        details_window = tk.Toplevel(self)
        details_window.title("Thông tin chi tiết môn học")
        details_window.geometry("800x400")

        frame = tk.Frame(details_window)
        frame.pack(fill=tk.BOTH, expand=True)

        details_tree = ttk.Treeview(frame, columns=("MonHoc", "Lop", "VangCoPhep", "VangKhongPhep", "TyLeVang"),
                                     show='headings')
        details_tree.heading("MonHoc", text="Môn học")
        details_tree.heading("Lop", text="Lớp")
        details_tree.heading("VangCoPhep", text="Vắng có phép")
        details_tree.heading("VangKhongPhep", text="Vắng không phép")
        details_tree.heading("TyLeVang", text="Tỷ lệ vắng")

        details_tree.column("MonHoc", width=200)
        details_tree.column("Lop", width=75)
        details_tree.column("VangCoPhep", width=150)
        details_tree.column("VangKhongPhep", width=150)
        details_tree.column("TyLeVang", width=150)

        details_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tree_scroll = tk.Scrollbar(frame, orient=tk.VERTICAL, command=details_tree.yview)
        details_tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        ngay_nghi_frame = tk.Frame(details_window)
        ngay_nghi_frame.pack(fill=tk.BOTH, expand=True)

        ngay_nghi_text = tk.Text(ngay_nghi_frame, wrap=tk.WORD, height=10)
        ngay_nghi_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        ngay_nghi_scroll = tk.Scrollbar(ngay_nghi_frame, orient=tk.VERTICAL, command=ngay_nghi_text.yview)
        ngay_nghi_text.configure(yscrollcommand=ngay_nghi_scroll.set)
        ngay_nghi_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        if student_details and "MonHocs" in student_details:
            for mon_hoc, details in student_details["MonHocs"].items():
                details_tree.insert('', tk.END, values=(
                    mon_hoc,
                    details["Lop"],
                    details["VangCoPhep"],
                    details["VangKhongPhep"],
                    details["TyLeVang"]
                ))
        else:
            details_tree.insert('', tk.END, values=("Không có thông tin chi tiết cho sinh viên này.", "", "", ""))
            ngay_nghi_text.insert(tk.END, "Không có thông tin ngày nghỉ cho sinh viên này.")

        ngay_nghi_text.config(state=tk.DISABLED)


        def on_tree_select(event):
            print("jaajaja")
            selected_item = details_tree.selection()
            if selected_item:
                index = selected_item[0]
                mon_hoc = details_tree.item(index)['values'][0]
                details = student_details["MonHocs"].get(mon_hoc)
                lop = details_tree.item(index)['values'][1]

                ngay_nghi_text.config(state=tk.NORMAL)
                ngay_nghi_text.delete(1.0, tk.END)

                if details:
                    ngay_nghi_text.insert(tk.END, f"Môn: {mon_hoc}\n")
                    ngay_nghi_text.insert(tk.END, f"Lớp: {lop}\n")
                    ngay_nghi_text.insert(tk.END, "Ngày nghỉ: \n" + "\n".join(
                        ngay['Ngay'] for ngay in details["NgayNghi"]) + "\n\n")
                else:
                    ngay_nghi_text.insert(tk.END, "Không có thông tin ngày nghỉ cho môn này.")

                    ngay_nghi_text.config(state=tk.DISABLED)

        details_tree.bind("<<TreeviewSelect>>", on_tree_select)




    def load_students(self):
        # Xóa các dòng cũ
        for row in self.tree.get_children():
            self.tree.delete(row)

            # Tải dữ liệu sinh viên từ cơ sở dữ liệu
        conn = sqlite3.connect('student_management.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM students")
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=(
                row[1],row[2],row[3]
            ))
        conn.close()

    def search_student(self):
        student_id = self.search_id_entry.get().strip()
        student_name = self.search_name_entry.get().strip().lower()

        # Xóa các dòng cũ
        for row in self.tree.get_children():
          self.tree.delete(row)

        # Tìm kiếm trong cơ sở dữ liệu
        conn = sqlite3.connect('student_management.db')
        cursor = conn.cursor()
        
        # Câu truy vấn tìm kiếm
        if student_id and student_name:
            cursor.execute("SELECT * FROM students WHERE student_id=? OR (lower(last_name) || ' ' || lower(first_name)) LIKE ?", (student_id, f'%{student_name}%'))
        elif student_id:
            cursor.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
        elif student_name:
            cursor.execute("SELECT * FROM students WHERE (lower(last_name) || ' ' || lower(first_name)) LIKE ?", (f'%{student_name}%',))
        else:
            self.load_students()  # Nếu không có gì để tìm kiếm, tải lại danh sách

        # Thêm kết quả tìm kiếm vào bảng
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()

    def show_add_student_form(self):
        self.student_form("Thêm sinh viên", None)

    def show_edit_student_form(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn sinh viên để sửa.")
            return

        item = self.tree.item(selected_item)
        student_data = item['values']

        self.student_form("Sửa sinh viên", student_data)

    def student_form(self, title, student_data):
        self.form_window = tk.Toplevel(self)
        self.form_window.title(title)
        self.form_window.minsize(400, 300)  # Kích thước tối thiểu lớn hơn

        tk.Label(self.form_window, text="Mã sinh viên:").grid(row=0, column=0, padx=5, pady=5)
        student_id_entry = tk.Entry(self.form_window)
        student_id_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self.form_window, text="Họ đệm:").grid(row=1, column=0, padx=5, pady=5)
        last_name_entry = tk.Entry(self.form_window)
        last_name_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(self.form_window, text="Tên:").grid(row=2, column=0, padx=5, pady=5)
        first_name_entry = tk.Entry(self.form_window)
        first_name_entry.grid(row=2, column=1, padx=5, pady=5)

        # Combobox chọn Lớp
        tk.Label(self.form_window, text="Lớp:").grid(row=3, column=0, padx=5, pady=5)
        self.class_combobox = ttk.Combobox(self.form_window)
        self.class_combobox.grid(row=3, column=1, padx=5, pady=5)

        # Combobox chọn Tên môn học
        tk.Label(self.form_window, text="Tên môn học:").grid(row=4, column=0, padx=5, pady=5)
        self.course_combobox = ttk.Combobox(self.form_window)
        self.course_combobox.grid(row=4, column=1, padx=5, pady=5)

        # Sự kiện thay đổi lớp học
        self.class_combobox.bind('<<ComboboxSelected>>', self.update_subjects)

        # Sự kiện thay đổi môn học
        self.course_combobox.bind('<<ComboboxSelected>>', self.update_classes)

        # Tải dữ liệu lớp học và môn học
        classes, subjects = self.load_classes_and_subjects()
        self.class_combobox['values'] = classes
        self.course_combobox['values'] = subjects

        if student_data:
            # Điền dữ liệu hiện có nếu có
            student_id_entry.insert(0, student_data[0])
            last_name_entry.insert(0, student_data[1])
            first_name_entry.insert(0, student_data[2])
            self.class_combobox.set(student_data[3])
            self.course_combobox.set(student_data[4])

        #Đặt lại hàng cho nút Lưu
        self.save_button = tk.Button(self.form_window, text="Lưu", command=lambda: self.save_student(
            student_id_entry.get(),
            last_name_entry.get(),
            first_name_entry.get(),
            self.class_combobox.get(),
            self.course_combobox.get(),
            student_data
        ))
        self.save_button.grid(row=5, column=0, columnspan=2, pady=10)  # Dùng grid và cột span để căn giữa

    def update_classes(self, event):
        selected_subject = self.course_combobox.get()
        classes = self.load_classes_for_subject(selected_subject)
        self.class_combobox['values'] = classes
        if classes:
            self.class_combobox.set(classes[0])  # Đặt giá trị mặc định

    def load_classes_for_subject(self, subject_name):
        # Truy vấn danh sách lớp học dựa trên môn học
        self.conn = sqlite3.connect('student_management.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT DISTINCT Lop FROM courses WHERE MonHoc = ?", (subject_name,))
        classes = [row[0] for row in self.cursor.fetchall()]
        self.conn.close()  # Đóng kết nối
        return classes

    def update_subjects(self, event):
        selected_class = self.class_combobox.get()
        subjects = self.load_subjects(selected_class)
        self.course_combobox['values'] = subjects
        if subjects:
            self.course_combobox.set(subjects[0])  # Đặt giá trị mặc định

    def load_subjects(self, class_name):
        # Truy vấn danh sách môn học dựa trên lớp học
        self.conn = sqlite3.connect('student_management.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT DISTINCT MonHoc FROM courses WHERE Lop = ?", (class_name,))
        subjects = [row[0] for row in self.cursor.fetchall()]
        self.conn.close()  # Đóng kết nối
        return subjects

    def load_classes_and_subjects(self):
        # Kết nối đến cơ sở dữ liệu
        self.conn = sqlite3.connect('student_management.db')
        self.cursor = self.conn.cursor()

        # Lấy danh sách lớp
        self.cursor.execute("SELECT DISTINCT Lop FROM courses")
        classes = [row[0] for row in self.cursor.fetchall()]
        #print("classes: " + str(classes))  # In ra để kiểm tra

        # Lấy danh sách tên môn học
        self.cursor.execute("SELECT DISTINCT MonHoc FROM courses")
        courses = [row[0] for row in self.cursor.fetchall()]
        #print("courses: " + str(courses))  # In ra để kiểm tra

        # Đóng kết nối với cơ sở dữ liệu
        self.conn.close()

        # Trả về danh sách lớp và môn học
        return classes, courses

    def save_student(self, student_id, last_name, first_name, class_name, subject_name, student_data=None):
        try:
            self.conn = sqlite3.connect('student_management.db')
            self.cursor = self.conn.cursor()
            # Bước 1: Lưu thông tin sinh viên vào bảng students
            if student_data is None:  # Thêm mới sinh viên
                self.cursor.execute("""
                   INSERT INTO students (MSSV, HoDem, Ten)
                   VALUES (?, ?, ?)
                   """, (student_id, last_name, first_name))
                self.conn.commit()
                student_id_db = self.cursor.lastrowid  # Lấy ID của sinh viên vừa thêm
                print("Mon hoc: "+subject_name)
                print("Lop: " + class_name)
                # Bước 2: Lưu thông tin khóa học vào bảng student_courses
                self.cursor.execute("SELECT ID FROM courses WHERE MonHoc = ? AND Lop = ?", (subject_name, class_name))
                course = self.cursor.fetchone()

                if course:
                    course_id = course[0]
                    self.cursor.execute("""
                       INSERT INTO student_courses (StudentID, CourseID)
                       VALUES (?, ?)
                       """, (student_id_db, course_id))
                    self.conn.commit()
                    messagebox.showinfo("Thông báo", "Thêm sinh viên thành công!")
                else:
                    messagebox.showerror("Lỗi", "Không tìm thấy khóa học với tên và lớp đã cho.")

            else:  # Cập nhật thông tin sinh viên
                student_id_db = student_data[0]  # ID sinh viên hiện tại
                self.cursor.execute("""
                   UPDATE students
                   SET MSSV = ?, HoDem = ?, Ten = ?
                   WHERE ID = ?
                   """, (student_id, last_name, first_name, student_id_db))
                self.conn.commit()

                messagebox.showinfo("Thông báo", "Cập nhật thông tin sinh viên thành công!")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")
        finally:
            # Đóng cửa sổ sau khi lưu
            self.form_window.destroy()
    
    def save_student_from_file(self, student_id, last_name, first_name,  co_phep, khong_phep, ti_le, id_course, p1, p2, p3, p4, p5, p6):
        #if (student_id and last_name and first_name):
            conn = sqlite3.connect('student_management.db')
            cursor = conn.cursor()

            cursor.execute("INSERT INTO students (MSSV, HoDem, Ten) VALUES (?, ?, ?)", 
                               (student_id, last_name, first_name))

            conn.commit()
            last_inserted_id = cursor.lastrowid
            cursor.execute("""
                                INSERT INTO student_courses (StudentID, CourseID, VangCoPhep, VangKhongPhep, TyLeVang)
                                VALUES (?, ?, ?, ?, ?)
                            """, (last_inserted_id, id_course, co_phep, khong_phep, ti_le))
            conn.commit()
            id_course_student = cursor.lastrowid
            co_phep = 1
            ko_phep = 0
            # print("P1: "+str(p1))
            # print("P2: " + str(p2))
            # print("P3: " + str(p3))
            # print("P4: " + str(p4))
            # print("P5: " + str(p5))
            # print("P6: " + str(p6))
            if p1 == 'P':
                #co_phep = co_phep + 1
                cursor.execute("""
                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                VALUES (?, ?, ?)
                                            """, (id_course_student, "11/06/2024", co_phep))
                conn.commit()
            elif p1 == 'K':
                #ko_phep = ko_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "11/06/2024", ko_phep))
                conn.commit()
            if p2 == 'P':
                #co_phep = co_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "18/06/2024", co_phep))
                conn.commit()
            elif p2== 'K':
                #ko_phep = ko_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "18/06/2024", ko_phep))
                conn.commit()
            if p3 == 'P':
                #co_phep = co_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "25/06/2024", co_phep))
                conn.commit()
            elif p3 == 'K':
                #ko_phep = ko_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "25/06/2024", ko_phep))
                conn.commit()
            if p4 == 'P':
                #co_phep = co_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "02/07/2024", co_phep))
                conn.commit()
            elif p4=='K':
                #ko_phep = ko_phep + 1
                cursor.execute("""
                                                                INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                VALUES (?, ?, ?)
                                                            """, (id_course_student, "02/07/2024", ko_phep))
                conn.commit()
            if p5 == 'P':
                    #co_phep = co_phep + 1
                    cursor.execute("""
                                                                    INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                    VALUES (?, ?, ?)
                                                                """, (id_course_student, "09/07/2024", co_phep))
                    conn.commit()
            elif p5 =='K':
                    #ko_phep = ko_phep + 1
                    cursor.execute("""
                                                                    INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                    VALUES (?, ?, ?)
                                                                """, (id_course_student, "09/07/2024", ko_phep))
                    conn.commit()
            if p6 == 'P':
                    #co_phep = co_phep + 1
                    cursor.execute("""
                                                                    INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                    VALUES (?, ?, ?)
                                                                """, (id_course_student, "23/07/2024", co_phep))
                    conn.commit()
            elif p6=='K':
                    #ko_phep = ko_phep + 1
                    cursor.execute("""
                                                                    INSERT INTO absences (StudentCourseID, NgayNghi, CoPhep)
                                                                    VALUES (?, ?, ?)
                                                                """, (id_course_student, "23/07/2024", ko_phep))
                    conn.commit()
            conn.close()
            self.load_students()
        # else:
        #     messagebox.showerror("Lỗi", "Vui lòng nhập thông tin hợp lệ.")

    def save_course_from_file(self, dot, monhoc, lop):
            conn = sqlite3.connect('student_management.db')
            cursor = conn.cursor()

            cursor.execute("INSERT INTO courses (MonHoc, Dot, Lop) VALUES (?, ?, ?)", 
                                (monhoc, dot, lop))
            last_inserted_id = cursor.lastrowid
            conn.commit()
            conn.close()
            self.load_students()
            return last_inserted_id
          

    def delete_student(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn sinh viên để xóa.")
            return

        student_id = self.tree.item(selected_item)['values'][0]
        conn = sqlite3.connect('student_management.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM students WHERE MSSV=?", (student_id,))
        conn.commit()
        conn.close()
        self.load_students()

    # def import_data(self):
    #   file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    #   if file_path:
    #       # Display a message to confirm file selection
    #       messagebox.showinfo("Thông báo", f"Đã chọn file: {file_path}")
    #
    #       # Check file extension to determine the library to use
    #       if file_path.endswith('.xlsx'):
    #           # Open the workbook using openpyxl for .xlsx files
    #           workbook = openpyxl.load_workbook(file_path)
    #           sheet = workbook.active  # Get the first sheet
    #
    #           # Read values from specific cells
    #           dot = sheet['C6'].value  # A6 in Excel corresponds to 'C6'
    #           monhoc = sheet['C9'].value  # A9 corresponds to 'C9'
    #           lop = sheet['C10'].value  # A10 corresponds to 'C10'
    #
    #           # Print values
    #           print(f"Giá trị A6 (dot): {dot}")
    #           print(f"Giá trị A9 (monhoc): {monhoc}")
    #           print(f"Giá trị A10 (lop): {lop}")
    #
    #           id_course = self.save_course_from_file(dot, monhoc, lop)
    #           # Loop through rows 14 to 60 and columns B to AB (index 1 to 27)
    #           for row in range(14, 61):
    #               row_data = []
    #               for col in range(2, 29):  # Columns B to AB
    #                   cell_value = sheet.cell(row=row, column=col).value
    #                   row_data.append(cell_value)
    #
    #               if len(row_data) >= 3:  # Ensure enough data is available in the row
    #                   student_id = row_data[0]
    #                   last_name = row_data[1]
    #                   first_name = row_data[2]
    #                   co_phep = row_data[23]
    #                   khong_phep = row_data[24]
    #                   ti_le = row_data[24]
    #                   p1 = row_data[5]
    #                   p2 = row_data[8]
    #                   p3 = row_data[11]
    #                   p4 = row_data[14]
    #                   p5 = row_data[17]
    #                   p6 = row_data[20]
    #                   self.save_student_from_file(student_id, last_name, first_name, co_phep, khong_phep, ti_le, id_course, p1, p2, p3, p4, p5, p6)
    #                   #print(f"Dòng {row}: {row_data}")
    #               else:
    #                   print(f"Dòng {row} không đủ dữ liệu.")
    #       else:
    #           messagebox.showerror("Lỗi", "Vui lòng chọn tệp Excel .xlsx!")
    def import_data(self):
        # Cho phép chọn nhiều file
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx;*.xls")])

        if file_paths:
            # Hiển thị thông báo xác nhận đã chọn file
            messagebox.showinfo("Thông báo", f"Đã chọn {len(file_paths)} file(s): {', '.join(file_paths)}")

            for file_path in file_paths:
                # Kiểm tra phần mở rộng file để xác định thư viện cần sử dụng
                if file_path.endswith('.xlsx'):
                    # Mở workbook sử dụng openpyxl cho file .xlsx
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active  # Lấy sheet đầu tiên

                    # Đọc giá trị từ các ô cụ thể
                    dot = sheet['C6'].value  # A6 trong Excel tương ứng với 'C6'
                    monhoc = sheet['C9'].value  # A9 tương ứng với 'C9'
                    lop = sheet['C10'].value  # A10 tương ứng với 'C10'

                    # In giá trị
                    print(f"Giá trị C6 (dot): {dot}")
                    print(f"Giá trị C9 (monhoc): {monhoc}")
                    print(f"Giá trị C10 (lop): {lop}")

                    id_course = self.save_course_from_file(dot, monhoc, lop)

                    # Vòng lặp qua các hàng từ 14 đến 60 và các cột từ B đến AB (index 1 đến 27)
                    for row in range(14, 61):
                        row_data = []
                        for col in range(2, 29):  # Các cột từ B đến AB
                            cell_value = sheet.cell(row=row, column=col).value
                            row_data.append(cell_value)

                        if len(row_data) >= 3:  # Đảm bảo đủ dữ liệu trong hàng
                            student_id = row_data[0]
                            last_name = row_data[1]
                            first_name = row_data[2]
                            co_phep = row_data[23]
                            khong_phep = row_data[24]
                            ti_le = row_data[26]  # Có thể bạn muốn sử dụng index 25 cho tỷ lệ vắng
                            #print("Ti le: "+ti_le)
                            p1 = row_data[5]
                            p2 = row_data[8]
                            p3 = row_data[11]
                            p4 = row_data[14]
                            p5 = row_data[17]
                            p6 = row_data[20]
                            self.save_student_from_file(student_id, last_name, first_name, co_phep, khong_phep, ti_le,
                                                        id_course, p1, p2, p3, p4, p5, p6)
                        else:
                            print(f"Dòng {row} không đủ dữ liệu.")
                else:
                    messagebox.showerror("Lỗi", "Vui lòng chọn tệp Excel .xlsx!")


if __name__ == "__main__":
    init_db()
    app = StudentManagementApp()
    app.mainloop()